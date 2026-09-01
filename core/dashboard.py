"""
RFP Dashboard — Flask web UI.
Entry point: run_dashboard.py  →  http://localhost:5000
"""

import csv
import io
import os
import re
import subprocess
import zipfile
from pathlib import Path
from datetime import date, datetime

from flask import Flask, render_template_string, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.utils import ROOT_DIR, LOGS_DIR

CSV_FILE = ROOT_DIR / "data" / "rfp_tracker.csv"
LOG_FILE  = LOGS_DIR / "rfp_agent.log"

CSV_FIELDS = [
    "date", "title", "source", "url", "file", "keywords_matched",
    "status",
    "approved_by", "approved_date",
    "applied_by",  "applied_date",
    "done_date",
    "notes",
]
VALID_STATUSES = {"new", "approved", "applied", "done", "closed"}

app = Flask(__name__)


# ── CSV helpers ───────────────────────────────────────────────────────────────

def load_rfps() -> list:
    if not CSV_FILE.exists():
        return []
    with open(CSV_FILE, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        for fld in CSV_FIELDS:
            r.setdefault(fld, "")
    return list(reversed(rows))


def _update_csv_row(url: str, updates: dict) -> bool:
    """Update any fields on the row matching `url` and rewrite the CSV."""
    if not CSV_FILE.exists():
        return False
    with open(CSV_FILE, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    found = False
    for row in rows:
        if row.get("url") == url:
            for k, v in updates.items():
                row[k] = v
            found = True
            break
    if not found:
        return False
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in CSV_FIELDS})
    return True


def get_stats(rows: list) -> dict:
    today   = date.today().strftime("%Y-%m-%d")
    sources = sorted({r.get("source", "") for r in rows if r.get("source")})
    return {
        "total":    len(rows),
        "today":    sum(1 for r in rows if r.get("date", "").startswith(today)),
        "approved": sum(1 for r in rows if r.get("status") == "approved"),
        "applied":  sum(1 for r in rows if r.get("status") == "applied"),
        "done":     sum(1 for r in rows if r.get("status") == "done"),
        "sources":  len(sources),
        "source_list": sources,
    }


def get_last_run() -> str:
    if not LOG_FILE.exists():
        return "Never"
    for line in reversed(LOG_FILE.read_text(encoding="utf-8", errors="ignore").splitlines()):
        if "RFP Agent" in line:
            return line[:19]
    return "—"


# ── Excel export ──────────────────────────────────────────────────────────────

def _build_excel(rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RFP Tracker"

    headers = [
        "Fetched Date", "Title", "Source", "URL", "File Path", "Keywords",
        "Status",
        "Approved By", "Approved Date",
        "Applied By",  "Applied Date",
        "Done Date",   "Notes",
    ]
    keys = CSV_FIELDS

    hdr_fill = PatternFill("solid", fgColor="1A1A2E")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    status_fills = {
        "approved": PatternFill("solid", fgColor="D1E7DD"),
        "applied":  PatternFill("solid", fgColor="CCE5FF"),
        "done":     PatternFill("solid", fgColor="D1ECF1"),
        "closed":   PatternFill("solid", fgColor="F8D7DA"),
    }
    status_labels = {
        "approved": "Approved", "applied": "Applied",
        "done": "Done", "closed": "Closed", "new": "New", "": "New",
    }

    for row_idx, r in enumerate(rows, 2):
        st   = (r.get("status") or "").strip()
        fill = status_fills.get(st)
        for col_idx, key in enumerate(keys, 1):
            val = r.get(key, "") or ""
            if key == "status":
                val = status_labels.get(val, val or "New")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(
                wrap_text=(key in ("title", "keywords_matched", "notes")),
                vertical="top",
            )
            if fill:
                cell.fill = fill
            if key == "url" and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[row_idx].height = 30

    col_widths = [18, 46, 20, 40, 38, 24, 11, 16, 16, 16, 16, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── PDF extraction helpers ────────────────────────────────────────────────────

def _pdf_text(pdf_bytes: bytes, max_pages: int = 12) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages[:max_pages])
    except Exception:
        return ""


def _is_toc_line(line: str) -> bool:
    return bool(re.search(r'\.{4,}\s*\d+\s*$', line.strip()))


def _split_paragraphs(text: str) -> list:
    paras, buf = [], []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            if buf:
                paras.append(" ".join(buf))
            buf = []
        elif not _is_toc_line(line):
            buf.append(line)
    if buf:
        paras.append(" ".join(buf))
    return [p for p in paras if len(p) > 20]


def extract_rfp_info(file_path: str) -> dict:
    path = Path(file_path)
    text = ""
    try:
        if path.suffix.lower() == ".zip":
            with zipfile.ZipFile(path) as z:
                candidates = sorted(
                    [f for f in z.namelist() if f.lower().endswith(".pdf")],
                    key=lambda n: (any(x in n.lower() for x in ("dfb","financial","boq")), n),
                )
                for name in candidates[:3]:
                    with z.open(name) as pf:
                        text += _pdf_text(pf.read()) + "\n"
                        if len(text) > 8000:
                            break
        elif path.suffix.lower() == ".pdf":
            text = _pdf_text(path.read_bytes(), max_pages=15)
    except Exception as e:
        return {"error": str(e), "raw_text": ""}

    info: dict = {"raw_text": text[:6000]}
    clean = "\n".join(l for l in text.splitlines() if not _is_toc_line(l))
    paras = _split_paragraphs(text)

    def _find(patterns, src=None):
        src = src or clean
        for pat in patterns:
            m = re.search(pat, src, re.IGNORECASE)
            if m:
                return m.group(1).strip()[:120]
        return ""

    if re.search(r"Request for Empanelment|RFE No\.", text, re.IGNORECASE):
        info["doc_type"] = "Request for Empanelment (RFE)"
    elif re.search(r"Request for Proposal|RFP No\.", text, re.IGNORECASE):
        info["doc_type"] = "Request for Proposal (RFP)"
    elif re.search(r"Request for Quotation|RFQ", text, re.IGNORECASE):
        info["doc_type"] = "Request for Quotation (RFQ)"
    elif re.search(r"Bid Document|GeM/\d{4}/B/", text, re.IGNORECASE):
        info["doc_type"] = "GeM Bid Document"
    else:
        info["doc_type"] = "Tender Document"

    info["ref_number"]      = _find([r"RFE No\.\s*[:\-]?\s*([^\n\r]{5,60})",r"RFP No\.\s*[:\-]?\s*([^\n\r]{5,60})",r"Bid Number\s*[:\-]?\s*([^\n\r]{5,40})",r"Tender No\.\s*[:\-]?\s*([^\n\r]{5,60})",r"(GEM/\d{4}/B/\d+)"])
    info["organization"]    = _find([r"Name of Organiz[a-z]+\s*[:\-]?\s*([^\n\r]{5,80})",r"Organisation Name\s*[:\-]?\s*([^\n\r]{5,80})",r"Procuring\s+Entity\s*[:\-]?\s*([^\n\r]{5,80})",r"Ministry\s*/\s*State Name\s*[:\-]?\s*([^\n\r]{5,80})"])
    info["ministry"]        = _find([r"Ministry of\s+([^\n\r]{5,80})",r"Department of\s+([^\n\r]{5,80})"])
    info["category"]        = _find([r"Service Category\s*[:\-]?\s*([^\n\r]{5,100})",r"Item Category\s*[:\-]?\s*([^\n\r]{5,100})",r"Empanelment Categories?\s*[:\-]?\s*([^\n\r]{5,100})"])
    info["deadline"]        = _find([r"Last Date.*?(?:Submission|Bid)\s*[:\-]?\s*([^\n\r]{5,60})",r"Bid End Date\s*/\s*Time\s*[:\-]?\s*([^\n\r]{5,60})",r"Closing Date\s*[:\-]?\s*([^\n\r]{5,60})",r"Deadline\s*[:\-]?\s*([^\n\r]{5,60})"])
    info["contract_period"] = _find([r"Contract.*?Period\s*[:\-]?\s*([^\n\r]{5,80})",r"Empanelment.*?Period\s*[:\-]?\s*([^\n\r]{5,80})",r"Duration\s*[:\-]?\s*([^\n\r]{5,80})"])
    info["emd"]             = _find([r"Earnest Money Deposit.*?(?:INR|Rs\.?)\s*([\d,]+)",r"EMD.*?(?:INR|Rs\.?)\s*([\d,]+)"])
    if info["emd"]:
        info["emd"] = "INR " + info["emd"]
    info["vendor_panel_size"] = _find([r"Vendor Panel Size\s*[:\-]?\s*([^\n\r]{5,80})",r"Number of Vendors?\s*[:\-]?\s*([^\n\r]{5,80})"])
    info["contact"]         = _find([r"([\w.+-]+@(?:nic\.in|nicsi\.nic\.in|gov\.in|[\w.-]+\.in))",r"Email\s*[:\-]?\s*([\w.@+-]+)"])

    _SUP = re.compile(r'\b(startup|MSE|MSME|DPIIT)\b', re.IGNORECASE)
    startup_paras = [p for p in paras if _SUP.search(p)]
    if startup_paras:
        info["startup_eligible"] = True
        info["startup_notes"] = " | ".join(p[:180] for p in sorted(startup_paras, key=len, reverse=True)[:3])
    else:
        info["startup_eligible"] = False
        info["startup_notes"] = ""

    scope_summary = ""
    scope_m = re.search(r'(?:^|\n)\s*(?:\d+[\.\d]*\s+)?SCOPE OF WORK\s*\n([\s\S]{1,2000})', clean, re.IGNORECASE)
    if scope_m:
        for line in scope_m.group(1).splitlines():
            line = line.strip()
            if re.match(r'^\d+[\.\d]*\s+[A-Z]', line) and len(line) < 60:
                break
            if len(line) >= 60:
                scope_summary = line[:350]; break
    if not scope_summary:
        for p in paras:
            if re.search(r'\b(objective|empanel|procure|provide|deploy|manage)\b', p, re.IGNORECASE) and len(p) >= 80:
                scope_summary = p[:350]; break
    info["scope_summary"] = scope_summary
    return info


# ── Template ──────────────────────────────────────────────────────────────────

TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RFP Dashboard</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css" rel="stylesheet">
<style>
  body { background:#f0f2f5; font-family:'Segoe UI',sans-serif; }
  .navbar { background:linear-gradient(135deg,#1a1a2e,#0f3460); }
  .stat-card { border:none; border-radius:14px; box-shadow:0 2px 12px rgba(0,0,0,.08); transition:transform .15s; }
  .stat-card:hover { transform:translateY(-3px); }
  .stat-icon { font-size:2rem; opacity:.8; }
  .table-card { border:none; border-radius:14px; box-shadow:0 2px 12px rgba(0,0,0,.08); overflow:hidden; }
  .table thead th { background:#1a1a2e; color:#fff; border:none; padding:11px 12px; font-weight:500; font-size:.83rem; }
  .table tbody tr:hover { background:#f4f6ff !important; }
  .filter-bar { background:white; border-radius:12px; padding:14px 18px;
                box-shadow:0 2px 8px rgba(0,0,0,.05); margin-bottom:18px; }
  .btn-open { font-size:.75rem; padding:2px 9px; }
  .kw-badge { font-size:.66rem; background:#e8f4fd; color:#0d6efd; border-radius:4px;
              padding:1px 5px; margin:1px; display:inline-block; }
  .title-cell { max-width:260px; }
  .no-data { text-align:center; padding:60px 20px; color:#888; }
  @keyframes spin { to { transform:rotate(360deg); } }
  .spinning { animation:spin .8s linear infinite; }

  /* ── Status badges ── */
  .st-badge { display:inline-block; font-size:.67rem; font-weight:700;
              padding:2px 8px; border-radius:10px; letter-spacing:.3px; white-space:nowrap; }
  .st-new      { background:#e9ecef; color:#495057; }
  .st-approved { background:#d1e7dd; color:#0a3622; }
  .st-applied  { background:#cce5ff; color:#00366e; }
  .st-done     { background:#d1ecf1; color:#0c5460; }
  .st-closed   { background:#f8d7da; color:#842029; }

  /* ── Row tints ── */
  tr[data-status="approved"] td { background:#f0fff4; }
  tr[data-status="applied"]  td { background:#f0f7ff; }
  tr[data-status="done"]     td { background:#f0fafa; }
  tr[data-status="closed"]   td { background:#fff5f5; opacity:.78; }

  /* ── Tracking info under badge ── */
  .track-info { font-size:.64rem; color:#6c757d; margin-top:3px; line-height:1.4; }
  .track-info strong { color:#333; }

  /* ── Status action buttons ── */
  .st-btn-grp { display:flex; gap:3px; justify-content:center; margin-bottom:5px; flex-wrap:nowrap; }
  .st-btn-grp .btn { font-size:.7rem; padding:2px 7px; }

  /* ── Detail panel ── */
  #detailPanel { position:fixed; top:0; right:-530px; width:510px; height:100vh;
    background:#fff; box-shadow:-4px 0 24px rgba(0,0,0,.15);
    transition:right .3s ease; z-index:1050; overflow-y:auto; }
  #detailPanel.open { right:0; }
  .panel-header { background:linear-gradient(135deg,#1a1a2e,#0f3460);
                  color:#fff; padding:16px 20px; position:sticky; top:0; z-index:1; }
  .info-row { display:flex; gap:8px; margin-bottom:9px; font-size:.86rem; }
  .info-label { color:#6c757d; min-width:118px; font-weight:500; flex-shrink:0; }
  .info-value { color:#1a1a2e; }
  .badge-startup    { background:#d4edda; color:#155724; border-radius:6px; padding:3px 10px; font-size:.76rem; font-weight:600; }
  .badge-no-startup { background:#f8d7da; color:#721c24; border-radius:6px; padding:3px 10px; font-size:.76rem; }
  .scope-box   { background:#f8f9fa; border-left:3px solid #0d6efd; padding:9px 13px;
                 border-radius:0 6px 6px 0; font-size:.84rem; color:#333; margin-top:4px; }
  .startup-box { background:#f0fff4; border-left:3px solid #28a745; padding:9px 13px;
                 border-radius:0 6px 6px 0; font-size:.81rem; color:#1a5e2a; margin-top:4px; line-height:1.5; }
  #panelOverlay { display:none; position:fixed; inset:0; background:rgba(0,0,0,.3); z-index:1040; }
  .loading-spinner { text-align:center; padding:60px 20px; color:#888; }

  /* ── Timeline ── */
  .timeline { border-left:2px solid #dee2e6; margin-left:10px; padding-left:16px; }
  .tl-item  { position:relative; margin-bottom:14px; }
  .tl-item::before { content:''; position:absolute; left:-21px; top:5px;
    width:10px; height:10px; border-radius:50%; background:#6c757d; border:2px solid #fff; }
  .tl-item.tl-approved::before { background:#198754; }
  .tl-item.tl-applied::before  { background:#0d6efd; }
  .tl-item.tl-done::before     { background:#0dcaf0; }
  .tl-item.tl-closed::before   { background:#dc3545; }
  .tl-label { font-size:.78rem; font-weight:700; color:#333; }
  .tl-meta  { font-size:.74rem; color:#6c757d; }
</style>
</head>
<body>

<nav class="navbar navbar-dark px-4 py-3 mb-4">
  <span class="navbar-brand fw-bold fs-5">
    <i class="bi bi-file-earmark-text me-2"></i>RFP Dashboard
  </span>
  <div class="d-flex align-items-center gap-2 flex-wrap">
    <span class="text-white-50" style="font-size:.8rem">
      Last run: <strong class="text-white">{{ last_run }}</strong>
    </span>
    <button class="btn btn-sm btn-light" onclick="runAgent()" id="runBtn">
      <i class="bi bi-play-circle me-1"></i>Run Agent Now
    </button>
    <a class="btn btn-sm btn-success" href="/export-excel">
      <i class="bi bi-file-earmark-excel me-1"></i>Export Excel
    </a>
    <button class="btn btn-sm btn-outline-light" onclick="location.reload()" title="Refresh">
      <i class="bi bi-arrow-clockwise"></i>
    </button>
  </div>
</nav>

<div class="container-fluid px-4">

  <!-- Stats -->
  <div class="row g-3 mb-4">
    <div class="col-6 col-md">
      <div class="card stat-card p-3">
        <div class="d-flex justify-content-between align-items-center">
          <div><div class="text-muted small">Total RFPs</div>
               <div class="fw-bold fs-3">{{ stats.total }}</div></div>
          <i class="bi bi-collection stat-icon text-primary"></i>
        </div>
      </div>
    </div>
    <div class="col-6 col-md">
      <div class="card stat-card p-3">
        <div class="d-flex justify-content-between align-items-center">
          <div><div class="text-muted small">Today</div>
               <div class="fw-bold fs-3 text-success">{{ stats.today }}</div></div>
          <i class="bi bi-calendar-check stat-icon text-success"></i>
        </div>
      </div>
    </div>
    <div class="col-6 col-md">
      <div class="card stat-card p-3">
        <div class="d-flex justify-content-between align-items-center">
          <div><div class="text-muted small">Approved</div>
               <div class="fw-bold fs-3" style="color:#0a3622">{{ stats.approved }}</div></div>
          <i class="bi bi-check-circle-fill stat-icon text-success"></i>
        </div>
      </div>
    </div>
    <div class="col-6 col-md">
      <div class="card stat-card p-3">
        <div class="d-flex justify-content-between align-items-center">
          <div><div class="text-muted small">Applied</div>
               <div class="fw-bold fs-3 text-primary">{{ stats.applied }}</div></div>
          <i class="bi bi-send-check stat-icon text-primary"></i>
        </div>
      </div>
    </div>
    <div class="col-6 col-md">
      <div class="card stat-card p-3">
        <div class="d-flex justify-content-between align-items-center">
          <div><div class="text-muted small">Done</div>
               <div class="fw-bold fs-3 text-info">{{ stats.done }}</div></div>
          <i class="bi bi-trophy stat-icon text-info"></i>
        </div>
      </div>
    </div>
    <div class="col-6 col-md">
      <div class="card stat-card p-3">
        <div class="d-flex justify-content-between align-items-center">
          <div><div class="text-muted small">Sources</div>
               <div class="fw-bold fs-3 text-secondary">{{ stats.sources }}</div></div>
          <i class="bi bi-globe stat-icon text-secondary"></i>
        </div>
      </div>
    </div>
  </div>

  <!-- Filters -->
  <div class="filter-bar d-flex flex-wrap gap-3 align-items-end">
    <div class="flex-grow-1" style="min-width:170px">
      <label class="form-label small text-muted mb-1">Search</label>
      <input type="text" id="searchInput" class="form-control form-control-sm"
             placeholder="Title, source, keywords…" oninput="filterTable()">
    </div>
    <div style="min-width:160px">
      <label class="form-label small text-muted mb-1">Source</label>
      <select id="sourceFilter" class="form-select form-select-sm" onchange="filterTable()">
        <option value="">All Sources</option>
        {% for s in stats.source_list %}<option value="{{ s }}">{{ s }}</option>{% endfor %}
      </select>
    </div>
    <div style="min-width:130px">
      <label class="form-label small text-muted mb-1">Status</label>
      <select id="statusFilter" class="form-select form-select-sm" onchange="filterTable()">
        <option value="">All</option>
        <option value="new">New</option>
        <option value="approved">Approved</option>
        <option value="applied">Applied</option>
        <option value="done">Done</option>
        <option value="closed">Closed</option>
      </select>
    </div>
    <div style="min-width:115px">
      <label class="form-label small text-muted mb-1">Date</label>
      <select id="dateFilter" class="form-select form-select-sm" onchange="filterTable()">
        <option value="">All Time</option>
        <option value="today">Today</option>
        <option value="week">Last 7 Days</option>
        <option value="month">Last 30 Days</option>
      </select>
    </div>
    <div style="min-width:100px">
      <label class="form-label small text-muted mb-1">File</label>
      <select id="fileFilter" class="form-select form-select-sm" onchange="filterTable()">
        <option value="">All</option>
        <option value="yes">Downloaded</option>
        <option value="no">No File</option>
      </select>
    </div>
    <button class="btn btn-sm btn-outline-secondary" onclick="clearFilters()">
      <i class="bi bi-x-circle me-1"></i>Clear
    </button>
    <span class="ms-auto text-muted small align-self-center" id="rowCount"></span>
  </div>

  <!-- Table -->
  <div class="card table-card mb-5">
    {% if rfps %}
    <div class="table-responsive">
      <table class="table table-hover mb-0" id="rfpTable">
        <thead>
          <tr>
            <th style="width:110px">Date</th>
            <th>Title</th>
            <th style="width:150px">Source</th>
            <th style="width:120px">Keywords</th>
            <th style="width:80px"  class="text-center">File</th>
            <th style="width:46px"  class="text-center">Link</th>
            <th style="width:160px" class="text-center">Status / Track</th>
          </tr>
        </thead>
        <tbody id="tableBody">
          {% for r in rfps %}
          {% set st = r.status if r.status in ('approved','applied','done','closed') else 'new' %}
          <tr data-source="{{ r.source|e }}"
              data-date="{{ r.date|e }}"
              data-file="{{ r.file|e }}"
              data-url="{{ r.url|e }}"
              data-title="{{ r.title|e }}"
              data-status="{{ st }}"
              data-approved-by="{{ r.approved_by|e }}"
              data-approved-date="{{ r.approved_date|e }}"
              data-applied-by="{{ r.applied_by|e }}"
              data-applied-date="{{ r.applied_date|e }}"
              data-done-date="{{ r.done_date|e }}"
              data-notes="{{ r.notes|e }}"
              onclick="rowClick(this)">
            <td class="text-muted small align-middle">{{ r.date[:16] if r.date else '' }}</td>
            <td class="title-cell align-middle" style="color:#1a1a2e">
              {{ r.title[:85] }}{% if r.title|length > 85 %}…{% endif %}
            </td>
            <td class="align-middle">
              <span class="badge text-bg-secondary" style="font-size:.69rem">{{ r.source[:28] }}</span>
            </td>
            <td class="align-middle">
              {% for kw in r.keywords_matched.split(',') if r.keywords_matched %}
              <span class="kw-badge">{{ kw.strip() }}</span>
              {% endfor %}
            </td>
            <td class="text-center align-middle" onclick="event.stopPropagation()">
              {% if r.file %}
              <button class="btn btn-success btn-open"
                      data-file="{{ r.file|e }}" data-url="{{ r.url|e }}" data-title="{{ r.title|e }}"
                      onclick="showDetails(this.dataset.file,this.dataset.url,this.dataset.title)">
                <i class="bi bi-file-earmark-text me-1"></i>View
              </button>
              {% else %}<span class="text-muted small">—</span>{% endif %}
            </td>
            <td class="text-center align-middle" onclick="event.stopPropagation()">
              {% if r.url %}
              <a href="{{ r.url }}" target="_blank" class="btn btn-outline-primary btn-open">
                <i class="bi bi-box-arrow-up-right"></i>
              </a>
              {% else %}<span class="text-muted">—</span>{% endif %}
            </td>

            <!-- ── Status + tracking cell ── -->
            <td class="text-center align-middle" onclick="event.stopPropagation()">
              <div class="st-btn-grp">
                <button title="Approve"
                        class="btn {{ 'btn-success' if st=='approved' else 'btn-outline-success' }}"
                        onclick="openModal(this.closest('tr'),'approved')">
                  <i class="bi bi-check-lg"></i>
                </button>
                <button title="Applied"
                        class="btn {{ 'btn-primary' if st=='applied' else 'btn-outline-primary' }}"
                        onclick="openModal(this.closest('tr'),'applied')">
                  <i class="bi bi-send"></i>
                </button>
                <button title="Done"
                        class="btn {{ 'btn-info' if st=='done' else 'btn-outline-info' }}"
                        onclick="openModal(this.closest('tr'),'done')">
                  <i class="bi bi-trophy"></i>
                </button>
                <button title="Close / Reject"
                        class="btn {{ 'btn-danger' if st=='closed' else 'btn-outline-danger' }}"
                        onclick="openModal(this.closest('tr'),'closed')">
                  <i class="bi bi-x-lg"></i>
                </button>
              </div>

              <!-- Current status badge -->
              <span class="st-badge st-{{ st }}">
                {% if st=='approved' %}✓ Approved
                {% elif st=='applied' %}✈ Applied
                {% elif st=='done'    %}★ Done
                {% elif st=='closed'  %}✗ Closed
                {% else %}New{% endif %}
              </span>

              <!-- Tracking info -->
              <div class="track-info">
                {% if r.approved_by %}
                  <div>✓ <strong>{{ r.approved_by }}</strong>
                  {% if r.approved_date %} · {{ r.approved_date }}{% endif %}</div>
                {% endif %}
                {% if r.applied_by %}
                  <div>✈ <strong>{{ r.applied_by }}</strong>
                  {% if r.applied_date %} · {{ r.applied_date }}{% endif %}</div>
                {% endif %}
                {% if r.done_date %}
                  <div>★ Done · {{ r.done_date }}</div>
                {% endif %}
                {% if r.notes %}
                  <div class="text-truncate" style="max-width:148px" title="{{ r.notes|e }}">
                    💬 {{ r.notes[:40] }}{% if r.notes|length > 40 %}…{% endif %}
                  </div>
                {% endif %}
              </div>
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
    {% else %}
    <div class="no-data">
      <i class="bi bi-inbox display-4 d-block mb-3 text-muted"></i>
      <h5 class="text-muted">No RFPs yet</h5>
      <p class="text-muted small">Click <strong>Run Agent Now</strong> to start fetching.</p>
    </div>
    {% endif %}
  </div>
</div>

<!-- ── Status Action Modal ──────────────────────────────────────────────────── -->
<div class="modal fade" id="statusModal" tabindex="-1">
  <div class="modal-dialog modal-sm">
    <div class="modal-content border-0 shadow">
      <div class="modal-header py-2" id="modalHeader">
        <h6 class="modal-title fw-bold" id="modalTitle">Update Status</h6>
        <button type="button" class="btn-close btn-sm" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body pb-2">
        <div id="nameGroup" class="mb-3">
          <label class="form-label small fw-bold mb-1" id="nameLabel">Your Name</label>
          <input type="text" id="modalName" class="form-control form-control-sm"
                 placeholder="e.g. Samira">
        </div>
        <div id="dateGroup" class="mb-3">
          <label class="form-label small fw-bold mb-1" id="dateLabel">Date</label>
          <input type="date" id="modalDate" class="form-control form-control-sm">
        </div>
        <div class="mb-2">
          <label class="form-label small fw-bold mb-1">
            Notes <span class="text-muted fw-normal">(optional)</span>
          </label>
          <textarea id="modalNotes" class="form-control form-control-sm" rows="2"
                    placeholder="Any remarks…"></textarea>
        </div>
      </div>
      <div class="modal-footer py-2">
        <button class="btn btn-sm btn-light" data-bs-dismiss="modal">Cancel</button>
        <button class="btn btn-sm" id="modalConfirmBtn" onclick="confirmStatus()">Confirm</button>
      </div>
    </div>
  </div>
</div>

<!-- ── Detail Panel ──────────────────────────────────────────────────────────── -->
<div id="panelOverlay" onclick="closePanel()"></div>
<div id="detailPanel">
  <div class="panel-header d-flex justify-content-between align-items-start">
    <div>
      <div class="small text-white-50 mb-1" id="panelDocType">—</div>
      <div class="fw-bold" id="panelRefNumber" style="font-size:.93rem">—</div>
    </div>
    <button class="btn btn-sm btn-outline-light ms-2 flex-shrink-0" onclick="closePanel()">
      <i class="bi bi-x-lg"></i>
    </button>
  </div>
  <div class="p-3">
    <div class="d-flex gap-2 mb-3">
      <button class="btn btn-success btn-sm" id="btnOpenFile" onclick="openFileDirect()">
        <i class="bi bi-folder2-open me-1"></i>Open File
      </button>
      <a class="btn btn-outline-primary btn-sm" id="btnOpenUrl" target="_blank" href="#">
        <i class="bi bi-box-arrow-up-right me-1"></i>View on Portal
      </a>
    </div>

    <!-- RFP content (spinner + data) -->
    <div class="loading-spinner" id="panelSpinner">
      <div class="spinner-border text-primary mb-3" role="status"></div>
      <div>Reading document…</div>
    </div>
    <div id="panelContent" style="display:none">
      <div class="mb-3">
        <div class="info-row"><span class="info-label"><i class="bi bi-building me-1"></i>Organisation</span><span class="info-value" id="piOrg">—</span></div>
        <div class="info-row"><span class="info-label"><i class="bi bi-diagram-3 me-1"></i>Ministry</span><span class="info-value" id="piMinistry">—</span></div>
        <div class="info-row"><span class="info-label"><i class="bi bi-tag me-1"></i>Category</span><span class="info-value" id="piCategory">—</span></div>
        <div class="info-row"><span class="info-label"><i class="bi bi-alarm me-1"></i>Deadline</span><span class="info-value fw-bold text-danger" id="piDeadline">—</span></div>
        <div class="info-row"><span class="info-label"><i class="bi bi-calendar3 me-1"></i>Contract</span><span class="info-value" id="piContract">—</span></div>
        <div class="info-row"><span class="info-label"><i class="bi bi-cash me-1"></i>EMD</span><span class="info-value" id="piEmd">—</span></div>
        <div class="info-row"><span class="info-label"><i class="bi bi-people me-1"></i>Vendor Slots</span><span class="info-value" id="piVendor">—</span></div>
        <div class="info-row"><span class="info-label"><i class="bi bi-envelope me-1"></i>Contact</span><span class="info-value" id="piContact">—</span></div>
      </div>
      <div class="mb-3">
        <div class="small text-muted fw-500 mb-1">Startup / MSE Eligibility</div>
        <div id="piStartupBadge"></div>
        <div class="startup-box mt-2" id="piStartupNotes" style="display:none"></div>
      </div>
      <div id="piScopeBlock" style="display:none">
        <div class="small text-muted fw-500 mb-1">Scope of Work</div>
        <div class="scope-box" id="piScope"></div>
      </div>
    </div>

    <!-- Tracking timeline — always visible, no doc needed -->
    <div id="panelTrackSection" class="mt-3" style="display:none">
      <hr class="my-3">
      <div class="small text-muted fw-bold mb-2"><i class="bi bi-clock-history me-1"></i>Tracking History</div>
      <div class="timeline" id="panelTimeline"></div>
      <div class="mt-3">
        <label class="form-label small fw-bold mb-1">Notes</label>
        <textarea id="panelNotesInput" class="form-control form-control-sm" rows="2"
                  placeholder="Add or edit notes…"></textarea>
        <button class="btn btn-sm btn-outline-secondary mt-2" onclick="saveNotes()">
          <i class="bi bi-save me-1"></i>Save Notes
        </button>
      </div>
    </div>
  </div>
</div>

<!-- Toast -->
<div class="toast-container position-fixed bottom-0 end-0 p-3">
  <div id="toast" class="toast border-0">
    <div class="d-flex">
      <div class="toast-body fw-500" id="toastText"></div>
      <button type="button" class="btn-close me-2 m-auto" data-bs-dismiss="toast"></button>
    </div>
  </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script>
const rows = document.querySelectorAll('#tableBody tr');
let currentFilePath = '';
let currentUrl      = '';

// ── Modal state ────────────────────────────────────────────────────────────────
let _mRow    = null;
let _mStatus = '';

const MODAL_CFG = {
  approved: {
    title:'Approve this RFP', headerBg:'#198754',
    nameLabel:'Approved By *', dateLbl:'Approval Date',
    showName:true, showDate:true,
    btnCls:'btn-success', btnTxt:'✓ Approve',
  },
  applied: {
    title:'Mark as Applied', headerBg:'#0d6efd',
    nameLabel:'Applied By *', dateLbl:'Application Date',
    showName:true, showDate:true,
    btnCls:'btn-primary', btnTxt:'✈ Mark Applied',
  },
  done: {
    title:'Mark as Done', headerBg:'#0dcaf0',
    nameLabel:'Completed By', dateLbl:'Completion Date',
    showName:true, showDate:true,
    btnCls:'btn-info', btnTxt:'★ Mark Done',
  },
  closed: {
    title:'Close / Reject', headerBg:'#dc3545',
    nameLabel:'Closed By', dateLbl:'',
    showName:true, showDate:false,
    btnCls:'btn-danger', btnTxt:'✗ Close',
  },
  new: {
    title:'Reset to New', headerBg:'#6c757d',
    showName:false, showDate:false,
    btnCls:'btn-secondary', btnTxt:'Reset',
  },
};

function openModal(row, status) {
  // Toggle: clicking active status resets to new
  const cur = row.dataset.status || 'new';
  const target = (cur === status) ? 'new' : status;
  _mRow = row; _mStatus = target;

  const cfg = MODAL_CFG[target] || MODAL_CFG.new;
  document.getElementById('modalTitle').textContent       = cfg.title;
  document.getElementById('modalHeader').style.background = cfg.headerBg || '#1a1a2e';
  document.getElementById('modalHeader').style.color      = '#fff';
  document.getElementById('nameGroup').style.display      = cfg.showName ? '' : 'none';
  document.getElementById('dateGroup').style.display      = cfg.showDate ? '' : 'none';
  document.getElementById('nameLabel').textContent        = cfg.nameLabel || 'Your Name';
  document.getElementById('dateLabel').textContent        = cfg.dateLbl  || 'Date';
  document.getElementById('modalDate').value              = new Date().toISOString().split('T')[0];
  document.getElementById('modalName').value              = '';
  document.getElementById('modalNotes').value             = row.dataset.notes || '';
  const btn = document.getElementById('modalConfirmBtn');
  btn.className = 'btn btn-sm ' + cfg.btnCls;
  btn.textContent = cfg.btnTxt;

  bootstrap.Modal.getOrCreateInstance(document.getElementById('statusModal')).show();
}

function confirmStatus() {
  const row    = _mRow;
  const status = _mStatus;
  const url    = row.dataset.url;
  const name   = document.getElementById('modalName').value.trim();
  const dt     = document.getElementById('modalDate').value;
  const notes  = document.getElementById('modalNotes').value.trim();

  bootstrap.Modal.getInstance(document.getElementById('statusModal')).hide();

  fetch('/set-status', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ url, status, name, action_date: dt, notes })
  })
  .then(r => r.json())
  .then(d => {
    if (!d.ok) { toast('Error: ' + d.message, 'danger'); return; }

    // Update data attributes on row
    row.dataset.status = status;
    row.dataset.notes  = notes;
    if (status === 'approved') { row.dataset.approvedBy = name; row.dataset.approvedDate = dt; }
    if (status === 'applied')  { row.dataset.appliedBy  = name; row.dataset.appliedDate  = dt; }
    if (status === 'done')     { row.dataset.doneDate   = dt; }

    // Refresh the status cell
    refreshStatusCell(row, status, name, dt, notes);

    const labels = { approved:'Approved ✓', applied:'Marked Applied ✈', done:'Marked Done ★', closed:'Closed ✗', new:'Reset to New' };
    toast(labels[status] || status, status==='approved'?'success':status==='applied'?'primary':status==='done'?'info':status==='closed'?'danger':'secondary');
  })
  .catch(() => toast('Network error','danger'));
}

const ST_ICON = { approved:'✓ Approved', applied:'✈ Applied', done:'★ Done', closed:'✗ Closed', new:'New' };
const ST_CLS  = { approved:'st-approved', applied:'st-applied', done:'st-done', closed:'st-closed', new:'st-new' };
const BTN_ACTIVE = {
  approved: ['btn-success',    'btn-outline-primary','btn-outline-info','btn-outline-danger'],
  applied:  ['btn-outline-success','btn-primary',    'btn-outline-info','btn-outline-danger'],
  done:     ['btn-outline-success','btn-outline-primary','btn-info',    'btn-outline-danger'],
  closed:   ['btn-outline-success','btn-outline-primary','btn-outline-info','btn-danger'],
  new:      ['btn-outline-success','btn-outline-primary','btn-outline-info','btn-outline-danger'],
};

function refreshStatusCell(row, status, name, dt, notes) {
  const td    = row.querySelector('td:last-child');
  const btns  = td.querySelectorAll('.st-btn-grp .btn');
  const pairs = [
    ['btn-success','btn-outline-success'],
    ['btn-primary','btn-outline-primary'],
    ['btn-info',   'btn-outline-info'],
    ['btn-danger', 'btn-outline-danger'],
  ];
  const active = BTN_ACTIVE[status] || BTN_ACTIVE.new;
  btns.forEach((b, i) => { b.classList.remove(...pairs[i]); b.classList.add(active[i]); });

  // Badge
  const badge = td.querySelector('.st-badge');
  badge.className = 'st-badge ' + (ST_CLS[status] || 'st-new');
  badge.textContent = ST_ICON[status] || 'New';

  // Tracking info
  const ti = td.querySelector('.track-info');
  let html = '';
  const ab  = row.dataset.approvedBy   || ''; const ad = row.dataset.approvedDate || '';
  const apb = row.dataset.appliedBy    || ''; const apd= row.dataset.appliedDate  || '';
  const dd  = row.dataset.doneDate     || '';
  const nt  = notes || row.dataset.notes || '';
  if (ab)  html += `<div>✓ <strong>${ab}</strong>${ad?' · '+ad:''}</div>`;
  if (apb) html += `<div>✈ <strong>${apb}</strong>${apd?' · '+apd:''}</div>`;
  if (dd)  html += `<div>★ Done · ${dd}</div>`;
  if (nt)  html += `<div class="text-truncate" style="max-width:148px" title="${nt}">💬 ${nt.slice(0,40)}${nt.length>40?'…':''}</div>`;
  ti.innerHTML = html;
}

// ── Filtering ─────────────────────────────────────────────────────────────────
function filterTable() {
  const q       = document.getElementById('searchInput').value.toLowerCase();
  const src     = document.getElementById('sourceFilter').value.toLowerCase();
  const statusF = document.getElementById('statusFilter').value;
  const dateF   = document.getElementById('dateFilter').value;
  const fileF   = document.getElementById('fileFilter').value;
  const now     = new Date();
  let vis = 0;
  rows.forEach(row => {
    const text    = row.textContent.toLowerCase();
    const rSrc    = (row.dataset.source||'').toLowerCase();
    const rDate   = row.dataset.date||'';
    const rFile   = (row.dataset.file||'').trim();
    const rStatus = row.dataset.status||'new';
    let show = true;
    if (q       && !text.includes(q))    show = false;
    if (src     && !rSrc.includes(src))  show = false;
    if (statusF && rStatus !== statusF)  show = false;
    if (fileF==='yes' && !rFile)         show = false;
    if (fileF==='no'  &&  rFile)         show = false;
    if (dateF && rDate) {
      const d = new Date(rDate);
      if (dateF==='today') { const t=new Date();t.setHours(0,0,0,0);if(d<t)show=false; }
      else if(dateF==='week') { if(d<new Date(now-7*864e5))show=false; }
      else if(dateF==='month'){ if(d<new Date(now-30*864e5))show=false; }
    }
    row.style.display = show?'':'none';
    if(show) vis++;
  });
  document.getElementById('rowCount').textContent = vis+' of '+rows.length+' RFPs';
}

function clearFilters() {
  ['searchInput','sourceFilter','statusFilter','dateFilter','fileFilter']
    .forEach(id=>document.getElementById(id).value='');
  filterTable();
}

// ── Row click → open detail panel ─────────────────────────────────────────────
function rowClick(row) {
  const file  = (row.dataset.file||'').trim();
  const url   = row.dataset.url||'';
  const title = row.dataset.title||'';
  currentUrl  = url;
  openPanel(file, url, title, row);
}

// ── Detail panel ──────────────────────────────────────────────────────────────
function openPanel(filePath, url, title, row) {
  currentFilePath = filePath;
  currentUrl      = url;
  const urlBtn = document.getElementById('btnOpenUrl');
  urlBtn.href = url||'#';
  urlBtn.style.display = url?'':'none';

  document.getElementById('panelRefNumber').textContent = title||'…';
  document.getElementById('panelDocType').textContent   = '';
  document.getElementById('panelSpinner').style.display  = filePath?'':'none';
  document.getElementById('panelContent').style.display  = 'none';

  // Show tracking section immediately
  buildTimeline(row);

  document.getElementById('detailPanel').classList.add('open');
  document.getElementById('panelOverlay').style.display = '';

  if (filePath) {
    fetch('/rfp-info', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({path: filePath})
    })
    .then(r=>r.json()).then(d=>populatePanel(d))
    .catch(e=>{
      document.getElementById('panelSpinner').style.display='none';
      document.getElementById('panelContent').innerHTML='<div class="text-danger p-2 small">Could not read document.</div>';
      document.getElementById('panelContent').style.display='';
    });
  }
}

function buildTimeline(row) {
  const section  = document.getElementById('panelTrackSection');
  const timeline = document.getElementById('panelTimeline');
  const status   = row ? row.dataset.status||'new' : 'new';
  const ab  = row?.dataset.approvedBy  || '';
  const ad  = row?.dataset.approvedDate|| '';
  const apb = row?.dataset.appliedBy   || '';
  const apd = row?.dataset.appliedDate || '';
  const dd  = row?.dataset.doneDate    || '';
  const nt  = row?.dataset.notes       || '';

  let html = '';
  html += tlItem('new','tl-new','Fetched / New',(row?.dataset.date||'').slice(0,16),'');
  if (ab || status==='approved'||status==='applied'||status==='done'||status==='closed')
    html += tlItem('approved','tl-approved','Approved', ad, ab);
  if (apb || status==='applied'||status==='done')
    html += tlItem('applied','tl-applied','Applied', apd, apb);
  if (dd || status==='done')
    html += tlItem('done','tl-done','Done', dd, '');
  if (status==='closed')
    html += tlItem('closed','tl-closed','Closed','','');

  timeline.innerHTML = html;
  document.getElementById('panelNotesInput').value = nt;
  section.style.display = '';
}

function tlItem(status, cls, label, dt, by) {
  return `<div class="tl-item ${cls}">
    <div class="tl-label">${label}</div>
    <div class="tl-meta">${by?'<strong>'+by+'</strong>':''} ${dt?'· '+dt:''}</div>
  </div>`;
}

function saveNotes() {
  const notes = document.getElementById('panelNotesInput').value.trim();
  if (!currentUrl) return;
  fetch('/set-status', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({url: currentUrl, status: null, notes, save_notes_only: true})
  })
  .then(r=>r.json()).then(d=>{
    toast(d.ok ? 'Notes saved' : 'Error: '+d.message, d.ok?'success':'danger');
    // Update the row's data attr
    rows.forEach(r=>{ if(r.dataset.url===currentUrl){ r.dataset.notes=notes; refreshStatusCell(r, r.dataset.status||'new', '', '', notes); }});
  });
}

function showDetails(filePath, url, title) {
  // Find the row for this file to get tracking data
  let targetRow = null;
  rows.forEach(r=>{ if(r.dataset.url===url) targetRow=r; });
  openPanel(filePath, url, title, targetRow);
}

function populatePanel(d) {
  document.getElementById('panelSpinner').style.display='none';
  document.getElementById('panelContent').style.display='';
  set('panelDocType',  d.doc_type||'');
  set('panelRefNumber', d.ref_number||document.getElementById('panelRefNumber').textContent);
  set('piOrg',       d.organization||'—');
  set('piMinistry',  d.ministry||'—');
  set('piCategory',  d.category||'—');
  set('piDeadline',  d.deadline||'—');
  set('piContract',  d.contract_period||'—');
  set('piEmd',       d.emd||'—');
  set('piVendor',    d.vendor_panel_size||'—');
  set('piContact',   d.contact||'—');
  const badge = document.getElementById('piStartupBadge');
  badge.innerHTML = d.startup_eligible
    ? '<span class="badge-startup"><i class="bi bi-check-circle-fill me-1"></i>Startup / MSE Eligible</span>'
    : '<span class="badge-no-startup"><i class="bi bi-x-circle me-1"></i>No startup clause detected</span>';
  const notesEl = document.getElementById('piStartupNotes');
  if (d.startup_notes) { notesEl.textContent=d.startup_notes; notesEl.style.display=''; }
  else notesEl.style.display='none';
  const scopeBlock = document.getElementById('piScopeBlock');
  if (d.scope_summary) { document.getElementById('piScope').textContent=d.scope_summary; scopeBlock.style.display=''; }
  else scopeBlock.style.display='none';
}

function set(id,val){const el=document.getElementById(id);if(el)el.textContent=val;}

function closePanel() {
  document.getElementById('detailPanel').classList.remove('open');
  document.getElementById('panelOverlay').style.display='none';
}

function openFileDirect() {
  if (!currentFilePath) return;
  fetch('/open-file',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({path:currentFilePath})})
  .then(r=>r.json()).then(d=>toast(d.message));
}

// ── Agent ─────────────────────────────────────────────────────────────────────
function runAgent() {
  const btn = document.getElementById('runBtn');
  btn.disabled=true;
  btn.innerHTML='<i class="bi bi-hourglass-split me-1 spinning"></i>Running…';
  toast('Agent started — fetching RFPs…');
  fetch('/run-agent',{method:'POST'})
  .then(r=>r.json()).then(d=>{
    toast(d.message);
    setTimeout(()=>location.reload(),8000);
  }).catch(()=>{btn.disabled=false;btn.innerHTML='<i class="bi bi-play-circle me-1"></i>Run Agent Now';});
}

function toast(msg, type) {
  const el = document.getElementById('toast');
  el.className = 'toast border-0 text-bg-'+(type||'primary');
  document.getElementById('toastText').textContent = msg;
  bootstrap.Toast.getOrCreateInstance(el,{delay:3500}).show();
}

filterTable();
setTimeout(()=>location.reload(), 300000);
</script>
</body>
</html>
"""


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    rfps     = load_rfps()
    stats    = get_stats(rfps)
    last_run = get_last_run()
    return render_template_string(TEMPLATE, rfps=rfps, stats=stats, last_run=last_run)


@app.route("/set-status", methods=["POST"])
def set_status():
    data   = request.get_json() or {}
    url    = data.get("url", "")
    status = data.get("status")
    name   = data.get("name", "").strip()
    dt     = data.get("action_date", "")
    notes  = data.get("notes", "").strip()
    notes_only = data.get("save_notes_only", False)

    if not url:
        return jsonify({"ok": False, "message": "No URL provided"})

    if notes_only:
        ok = _update_csv_row(url, {"notes": notes})
        return jsonify({"ok": ok, "message": "Notes saved" if ok else "RFP not found"})

    if status not in VALID_STATUSES:
        return jsonify({"ok": False, "message": "Invalid status"})

    now_str = datetime.now().strftime("%Y-%m-%d")
    updates = {"status": status, "notes": notes}

    if status == "approved":
        updates["approved_by"]   = name
        updates["approved_date"] = dt or now_str
    elif status == "applied":
        updates["applied_by"]   = name
        updates["applied_date"] = dt or now_str
    elif status == "done":
        updates["done_date"] = dt or now_str
        if name:
            updates["applied_by"] = updates.get("applied_by") or name
    elif status == "new":
        # Reset all tracking fields
        updates.update({
            "approved_by": "", "approved_date": "",
            "applied_by":  "", "applied_date":  "",
            "done_date":   "",
        })

    ok = _update_csv_row(url, updates)
    return jsonify({"ok": ok, "message": f"Status → {status}" if ok else "RFP not found"})


@app.route("/export-excel")
def export_excel():
    rows     = load_rfps()
    filename = f"rfp_tracker_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        io.BytesIO(_build_excel(rows)),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/rfp-info", methods=["POST"])
def rfp_info():
    path = (request.get_json() or {}).get("path", "")
    p = Path(path)
    if not p.is_absolute():
        p = ROOT_DIR / p
    if not p.exists():
        return jsonify({"error": f"File not found: {p.name}"})
    try:
        info = extract_rfp_info(str(p))
        info.pop("raw_text", None)
        return jsonify(info)
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route("/open-file", methods=["POST"])
def open_file():
    path = (request.get_json() or {}).get("path", "")
    p = Path(path)
    if not p.is_absolute():
        p = ROOT_DIR / p
    if not p.exists():
        return jsonify({"message": f"File not found: {p.name}"})
    try:
        os.startfile(str(p))
        return jsonify({"message": f"Opened: {p.name}"})
    except Exception as e:
        return jsonify({"message": f"Error: {e}"})


@app.route("/run-agent", methods=["POST"])
def run_agent():
    script = ROOT_DIR / "run_agent.py"
    py     = ROOT_DIR / ".venv" / "Scripts" / "python.exe"
    exe    = str(py) if py.exists() else "python"
    try:
        subprocess.Popen(
            [exe, str(script)],
            cwd=str(ROOT_DIR),
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )
        return jsonify({"message": "Agent running — page will refresh shortly."})
    except Exception as e:
        return jsonify({"message": f"Failed: {e}"})
